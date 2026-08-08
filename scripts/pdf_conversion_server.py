#!/usr/bin/env python3
"""
Self-hosted PDF -> Word / PDF -> Excel conversion service — a no-credentials
alternative to the Adobe PDF Services provider (see
lib/pdf-conversion/self-hosted-provider.ts, selected via
PDF_CONVERSION_PROVIDER=self-hosted). Everything runs on this VPS, so
there's no external API, no signup, and no per-conversion cost.

PDF -> DOCX uses pdf2docx, which reconstructs real paragraphs, basic
tables, and images rather than a flat text dump.

PDF -> XLSX uses pdfplumber's table detection to rebuild actual rows and
columns per page (one worksheet per detected table). If no table is
detected anywhere in the document, it falls back to one line of text per
row — the same floor the old client-side conversion had — rather than
producing an empty workbook.

Each cell is matched back to its bounding box in the source PDF, so the
sheet mirrors the page rather than imposing a house style: cell shading,
text color, bold/italic, font size and alignment are read out of the PDF
itself; merged regions are re-merged; row heights and column widths carry
over from the page geometry; and numbers become real numeric cells whose
format matches how they were already written (including Indian
lakh/crore grouping). Borders are drawn only where the PDF was actually
ruled, so an unruled document doesn't gain a grid it never had.

Run under the dedicated venv at /opt/pdfgenie-ml/venv (see whisper_server.py
for the sibling self-hosted service this matches).
"""
import os
import re
import tempfile

from flask import Flask, Response, jsonify, request
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pdf2docx import Converter
from pypdf import PdfReader, PdfWriter
import pdfplumber

PORT = int(os.environ.get("PDF_CONVERSION_PORT", "5007"))
DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

app = Flask(__name__)


@app.get("/health")
def health():
    return jsonify({"status": "ok"})


def _save_upload(pdf_file) -> str:
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        pdf_file.save(tmp.name)
        return tmp.name


def _decrypt_if_needed(input_path: str, password: str | None) -> tuple[str | None, tuple[str, int] | None]:
    """Returns (path_to_use, None) on success, or (None, (error_key, status)) on failure. path_to_use may be a new decrypted file, or the original path if the PDF wasn't encrypted."""
    reader = PdfReader(input_path)
    if not reader.is_encrypted:
        return input_path, None
    if not password:
        return None, ("password_required", 423)
    if reader.decrypt(password) == 0:
        return None, ("incorrect_password", 401)

    writer = PdfWriter()
    for page in reader.pages:
        writer.add_page(page)
    decrypted_path = input_path + ".decrypted.pdf"
    with open(decrypted_path, "wb") as f:
        writer.write(f)
    return decrypted_path, None


def _looks_like_a_real_table(table) -> bool:
    if not table or len(table) < 2:
        return False
    if not any(len(row) >= 2 for row in table):
        return False
    non_empty_cells = sum(1 for row in table for cell in row if cell and str(cell).strip())
    return non_empty_cells >= 4


def _find_tables_multi_strategy(page):
    """pdfplumber's default ("lines") strategy only finds tables with actual
    drawn ruling — most real-world documents (invoices, exports, reports)
    align columns with whitespace instead, so this also tries a "text"
    (alignment-based) pass when the ruled-line pass finds nothing usable.
    Returns (tables, ruled): Table objects (not just their extracted grids) so
    each cell's bounding box is available for formatting lookups, plus whether
    they came from real ruling — only then do borders and merges reflect
    something actually drawn in the PDF rather than something we inferred."""
    tables = [t for t in page.find_tables() if _looks_like_a_real_table(t.extract())]
    if tables:
        return tables, True

    text_settings = {"vertical_strategy": "text", "horizontal_strategy": "text"}
    tables = [t for t in page.find_tables(text_settings) if _looks_like_a_real_table(t.extract())]
    return tables, False


_THIN_SIDE = Side(style="thin", color="000000")
_CELL_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)
_HEADER_FILL = PatternFill(start_color="E7EAF3", end_color="E7EAF3", fill_type="solid")
_NUMERIC_RE = re.compile(r"^\(?-?\$?\s*\d[\d,]*(\.\d+)?%?\)?$")
# "1,08,519.00" — a group of exactly 2 digits before the final group of 3 is the
# Indian lakh/crore convention, which Excel's plain #,##0 would regroup wrongly.
_INDIAN_GROUPING_RE = re.compile(r"\d,\d{2},\d{3}(\.\d+)?$")
_INDIAN_NUMBER_FORMAT = "[>=10000000]##\\,##\\,##\\,##0.00;[>=100000]##\\,##\\,##0.00;##,##0.00"


def _color_to_hex(color):
    """PDF colors arrive as a gray scalar, or an RGB/CMYK sequence. Returns an
    'RRGGBB' string, or None when there's nothing usable to convert."""
    if color is None:
        return None
    if isinstance(color, (int, float)):
        v = max(0, min(255, int(round(float(color) * 255))))
        return f"{v:02X}{v:02X}{v:02X}"
    try:
        comps = [float(c) for c in color]
    except (TypeError, ValueError):
        return None
    if len(comps) == 1:
        v = max(0, min(255, int(round(comps[0] * 255))))
        return f"{v:02X}{v:02X}{v:02X}"
    if len(comps) == 3:
        return "".join(f"{max(0, min(255, int(round(c * 255)))):02X}" for c in comps)
    if len(comps) == 4:
        c, m, y, k = comps
        rgb_vals = (255 * (1 - min(1.0, c + k)), 255 * (1 - min(1.0, m + k)), 255 * (1 - min(1.0, y + k)))
        return "".join(f"{max(0, min(255, int(round(v)))):02X}" for v in rgb_vals)
    return None


def _fill_rects(page):
    """Filled (non-white) rectangles, smallest first so the innermost/most
    specific shading wins when boxes are nested."""
    out = []
    for r in page.rects:
        if not r.get("fill"):
            continue
        hex_color = _color_to_hex(r.get("non_stroking_color"))
        if not hex_color or hex_color == "FFFFFF":
            continue
        area = max(0.0, (r["x1"] - r["x0"])) * max(0.0, (r["bottom"] - r["top"]))
        out.append((area, r["x0"], r["top"], r["x1"], r["bottom"], hex_color))
    out.sort(key=lambda item: item[0])
    return out


def _fill_for_bbox(fills, bbox):
    """The shading actually painted behind a cell in the source PDF."""
    if not bbox or not fills:
        return None
    cx = (bbox[0] + bbox[2]) / 2
    cy = (bbox[1] + bbox[3]) / 2
    for _area, x0, top, x1, bottom, hex_color in fills:
        if x0 - 1 <= cx <= x1 + 1 and top - 1 <= cy <= bottom + 1:
            return hex_color
    return None


def _cell_chars(page_chars, bbox):
    """Characters from the page whose center point falls inside a cell's bounding box."""
    if not bbox:
        return []
    x0, top, x1, bottom = bbox
    out = []
    for c in page_chars:
        cx = (c["x0"] + c["x1"]) / 2
        cy = (c["top"] + c["bottom"]) / 2
        if x0 - 1 <= cx <= x1 + 1 and top - 1 <= cy <= bottom + 1:
            out.append(c)
    return out


def _is_bold(chars) -> bool:
    if not chars:
        return False
    bold_count = sum(1 for c in chars if "bold" in (c.get("fontname") or "").lower())
    return bold_count > len(chars) / 2


def _is_italic(chars) -> bool:
    if not chars:
        return False
    keys = ("italic", "oblique")
    count = sum(1 for c in chars if any(k in (c.get("fontname") or "").lower() for k in keys))
    return count > len(chars) / 2


def _text_color(chars):
    """Most common ink color among a cell's characters."""
    counts: dict[str, int] = {}
    for c in chars:
        hex_color = _color_to_hex(c.get("non_stroking_color"))
        if hex_color:
            counts[hex_color] = counts.get(hex_color, 0) + 1
    if not counts:
        return None
    return max(counts.items(), key=lambda kv: kv[1])[0]


def _font_size(chars):
    sizes = sorted(c["size"] for c in chars if c.get("size"))
    if not sizes:
        return None
    return max(6.0, min(36.0, round(sizes[len(sizes) // 2], 1)))


def _horizontal_alignment(chars, bbox):
    """Recovers left/center/right from where the glyphs actually sit in the cell."""
    if not chars or not bbox:
        return None
    text_x0 = min(c["x0"] for c in chars)
    text_x1 = max(c["x1"] for c in chars)
    left_gap = text_x0 - bbox[0]
    right_gap = bbox[2] - text_x1
    if left_gap < 0 or right_gap < 0:
        return None
    if right_gap < left_gap - 3:
        return "right"
    if left_gap > 4 and abs(left_gap - right_gap) <= 3:
        return "center"
    return "left"


def _parse_numeric_cell(text: str):
    """Recognizes plain/currency/percent numbers (incl. parenthesized negatives
    like accounting notation) and returns (value, is_currency, is_percent), or
    None if the text isn't a single number. Codes with a leading zero (invoice
    numbers, zip codes, IDs) are deliberately left as text — converting "007"
    to 7 would silently change what the cell displays."""
    t = text.strip()
    if not t or not _NUMERIC_RE.match(t):
        return None
    is_percent = t.endswith("%")
    is_currency = "$" in t
    negative = t.startswith("-") or (t.startswith("(") and t.endswith(")"))
    cleaned = t.strip("()").replace("$", "").replace("%", "").replace(",", "").lstrip("-")
    if not cleaned:
        return None
    int_part = cleaned.split(".")[0]
    if len(int_part) > 1 and int_part[0] == "0" and not is_currency and not is_percent:
        return None
    try:
        value = float(cleaned)
    except ValueError:
        return None
    if negative:
        value = -value
    if is_percent:
        value /= 100
    return value, is_currency, is_percent


def _number_format_for(text: str, is_currency: bool, is_percent: bool) -> str:
    """Mirrors how the number was already written in the PDF — a value shown as
    "39022.00" keeps its lack of separators, "1,08,519.00" keeps lakh grouping —
    so the converted sheet reads the same as the page it came from."""
    if is_percent:
        return "0.00%"
    stripped = text.strip().strip("()")
    if _INDIAN_GROUPING_RE.search(stripped):
        return _INDIAN_NUMBER_FORMAT
    grouped = "," in stripped
    decimals = "." in stripped
    base = ("#,##0" if grouped else "0") + (".00" if decimals else "")
    return f'"$"{base}' if is_currency else base


def _collect_cells(table):
    """Flattens a detected table into placed cells with real row/column spans.

    pdfplumber reports a merged region as one wide/tall bounding box plus
    `None` in the positions it swallowed, so spans can't be read off the grid
    directly. Instead, column indices are derived from the distinct cell edges
    across the whole table, which recovers a colspan from the box's own width
    and a rowspan from the same box reappearing on a later row."""
    col_starts = sorted({round(c[0], 1) for row in table.rows for c in row.cells if c})
    if not col_starts:
        return [], 0

    def col_of(x):
        return min(range(len(col_starts)), key=lambda i: abs(col_starts[i] - x))

    def col_end_of(x1):
        # Exclusive end: how many column starts sit strictly left of this edge.
        return max(1, sum(1 for s in col_starts if s < x1 - 1))

    placed: dict[tuple, dict] = {}
    data = table.extract()
    for row_index, (row_text, row_obj) in enumerate(zip(data, table.rows)):
        for i, bbox in enumerate(row_obj.cells):
            if bbox is None:
                continue
            key = tuple(round(v, 1) for v in bbox)
            existing = placed.get(key)
            if existing is not None:
                existing["row_end"] = row_index  # same box again → it spans rows
                continue
            start = col_of(bbox[0])
            placed[key] = {
                "row": row_index,
                "row_end": row_index,
                "col": start,
                "col_end": max(start, col_end_of(bbox[2]) - 1),
                "text": row_text[i] if i < len(row_text) else None,
                "bbox": bbox,
            }
    return list(placed.values()), len(col_starts)


def _style_and_write_table(ws, page, table, ruled: bool):
    """Writes one detected table into a worksheet, mirroring how the source PDF
    actually looks rather than imposing a house style: cell shading, text
    color, bold/italic and font size are read back out of the PDF, merged
    regions are re-merged, row heights and column widths are carried over in
    points, and numbers become real numeric cells with a matching format
    (including Indian lakh/crore grouping) instead of formatted strings.
    Borders are only drawn for tables that were actually ruled in the PDF."""
    cells, col_count = _collect_cells(table)
    if not cells:
        return

    page_chars = page.chars
    fills = _fill_rects(page)

    # A ruled table's blank rows are real layout, so they're kept — and merge
    # ranges depend on those row numbers staying put. The "text" strategy, by
    # contrast, invents blank rows between real ones (and never yields merges),
    # so there the rows are compacted.
    if ruled:
        row_map = {i: i for i in range(len(table.rows))}
    else:
        row_map = {}
        for i, row_text in enumerate(table.extract()):
            if any(c and str(c).strip() for c in row_text):
                row_map[i] = len(row_map)

    for item in cells:
        if item["row"] not in row_map:
            continue
        bbox = item["bbox"]
        raw = item["text"]
        value = "" if raw is None else str(raw).strip()
        chars = _cell_chars(page_chars, bbox)

        row = row_map[item["row"]] + 1
        row_end = row_map.get(item["row_end"], row_map[item["row"]])
        col = item["col"] + 1
        cell = ws.cell(row=row, column=col)

        numeric = _parse_numeric_cell(value)
        if numeric is not None:
            num_value, is_currency, is_percent = numeric
            cell.value = num_value
            cell.number_format = _number_format_for(value, is_currency, is_percent)
        else:
            cell.value = value.replace("\n", " ") if value else None

        align = _horizontal_alignment(chars, bbox) or ("right" if numeric is not None else "left")
        cell.alignment = Alignment(
            horizontal=align,
            vertical="center",
            wrap_text=len(value) > 60,
        )
        cell.font = Font(
            bold=_is_bold(chars),
            italic=_is_italic(chars),
            size=_font_size(chars),
            color=_text_color(chars),
        )

        fill_hex = _fill_for_bbox(fills, bbox)
        # Style every cell the merge will cover, not just the anchor, so the
        # shading and outline survive the merge.
        for r in range(row, row_end + 2):
            for c in range(col, item["col_end"] + 2):
                target = ws.cell(row=r, column=c)
                if ruled:
                    target.border = _CELL_BORDER
                if fill_hex:
                    target.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")

        if ruled and (row_end + 1 > row or item["col_end"] + 1 > col):
            ws.merge_cells(
                start_row=row,
                start_column=col,
                end_row=row_end + 1,
                end_column=item["col_end"] + 1,
            )

    # Excel row height is already in points, so PDF geometry transfers directly.
    for row_index, row_obj in enumerate(table.rows):
        if row_index not in row_map or not row_obj.bbox:
            continue
        height = row_obj.bbox[3] - row_obj.bbox[1]
        if height > 0:
            ws.row_dimensions[row_map[row_index] + 1].height = round(min(height, 409), 1)

    # Column width is measured in characters; ~7 points per character at the
    # default font is the conventional approximation.
    widths: dict[int, float] = {}
    for item in cells:
        if item["col_end"] != item["col"]:
            continue  # a merged span says nothing about any single column
        points = item["bbox"][2] - item["bbox"][0]
        widths[item["col"]] = max(widths.get(item["col"], 0.0), points / 7.0)
    for col_index in range(col_count):
        ws.column_dimensions[get_column_letter(col_index + 1)].width = round(
            max(3.0, min(widths.get(col_index, 10.0), 120.0)), 2
        )


@app.post("/convert-to-docx")
def convert_to_docx():
    if "file" not in request.files:
        return jsonify({"error": "No file provided."}), 400

    input_path = _save_upload(request.files["file"])
    password = request.form.get("password") or None
    cleanup_paths = [input_path]

    try:
        source_path, error = _decrypt_if_needed(input_path, password)
        if error:
            return jsonify({"error": error[0]}), error[1]
        if source_path != input_path:
            cleanup_paths.append(source_path)

        docx_path = input_path + ".out.docx"
        cleanup_paths.append(docx_path)

        cv = Converter(source_path)
        try:
            cv.convert(docx_path)
        finally:
            cv.close()

        with open(docx_path, "rb") as f:
            data = f.read()
        return Response(data, mimetype=DOCX_MIME)
    except Exception as exc:  # noqa: BLE001 — surfaced to the caller, not swallowed
        return jsonify({"error": str(exc)}), 500
    finally:
        for path in cleanup_paths:
            if path and os.path.exists(path):
                os.unlink(path)


@app.post("/convert-to-xlsx")
def convert_to_xlsx():
    if "file" not in request.files:
        return jsonify({"error": "No file provided."}), 400

    input_path = _save_upload(request.files["file"])
    password = request.form.get("password") or None
    cleanup_paths = [input_path]

    try:
        source_path, error = _decrypt_if_needed(input_path, password)
        if error:
            return jsonify({"error": error[0]}), error[1]
        if source_path != input_path:
            cleanup_paths.append(source_path)

        wb = Workbook()
        wb.remove(wb.active)
        sheet_count = 0
        used_names = set()

        def add_sheet(name: str):
            base = name[:31] or "Sheet"
            candidate = base
            i = 2
            while candidate in used_names:
                suffix = f" ({i})"
                candidate = base[: 31 - len(suffix)] + suffix
                i += 1
            used_names.add(candidate)
            return wb.create_sheet(title=candidate)

        with pdfplumber.open(source_path) as pdf:
            for page_index, page in enumerate(pdf.pages):
                tables, ruled = _find_tables_multi_strategy(page)
                for table_index, table in enumerate(tables):
                    sheet_count += 1
                    label = f"Page {page_index + 1}" if len(tables) == 1 else f"Page {page_index + 1} Table {table_index + 1}"
                    ws = add_sheet(label)
                    _style_and_write_table(ws, page, table, ruled)

        if sheet_count == 0:
            # No detectable table anywhere, even with the looser strategies
            # above — fall back to text, but still split each line into
            # columns wherever there's a run of 2+ spaces (the common
            # whitespace-alignment convention), rather than dumping the
            # whole line into one cell.
            ws = add_sheet("Text")
            header = ["Page", "Column 1", "Column 2", "Column 3", "Column 4", "Column 5"]
            ws.append(header)
            for col_index in range(len(header)):
                cell = ws.cell(row=1, column=col_index + 1)
                cell.font = Font(bold=True)
                cell.fill = _HEADER_FILL
                cell.border = _CELL_BORDER
                ws.column_dimensions[get_column_letter(col_index + 1)].width = 20
            with pdfplumber.open(source_path) as pdf:
                for page_index, page in enumerate(pdf.pages):
                    text = page.extract_text() or ""
                    for line in text.splitlines():
                        if line.strip():
                            cells = re.split(r" {2,}|\t", line.strip())
                            ws.append([page_index + 1, *cells])
            ws.freeze_panes = "A2"

        xlsx_path = input_path + ".out.xlsx"
        cleanup_paths.append(xlsx_path)
        wb.save(xlsx_path)

        with open(xlsx_path, "rb") as f:
            data = f.read()
        return Response(data, mimetype=XLSX_MIME)
    except Exception as exc:  # noqa: BLE001
        return jsonify({"error": str(exc)}), 500
    finally:
        for path in cleanup_paths:
            if path and os.path.exists(path):
                os.unlink(path)


if __name__ == "__main__":
    # threaded=True so one slow conversion doesn't block other requests —
    # still Flask's dev server (fine for an internal, localhost-only service
    # with modest concurrent load; never exposed to the internet).
    app.run(host="127.0.0.1", port=PORT, threaded=True)
